# core_views.py
from flask import Blueprint, jsonify, request, session, current_app, send_file
from flask_login import login_required, current_user
from models import (
    db,
    Credential,
    Account,
    Transaction,
    CustomCategory,
    CategoryRule,
    TransactionCategoryOverride,
    Budget
)
from plaid.model.item_remove_request import ItemRemoveRequest
from plaid.model.accounts_get_request import AccountsGetRequest
from plaid.model.item_webhook_update_request import ItemWebhookUpdateRequest
#from plaid.model.accounts_balance_get_request import AccountsBalanceGetRequest
from plaid.model.item_public_token_exchange_request import ItemPublicTokenExchangeRequest
from sqlalchemy import and_, or_, func, asc, desc, inspect, text
from sqlalchemy.orm import joinedload, aliased
import plaid
import json
import re
from datetime import datetime, date
from collections import defaultdict
import calendar
from io import StringIO, BytesIO
import csv
try:
    from openpyxl import Workbook
except ImportError:  # pragma: no cover - Excel export optional
    Workbook = None
from config import Config
from version import __version__ as VERSION
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from sqlalchemy.exc import OperationalError, IntegrityError
from uuid import uuid4

core = Blueprint('core', __name__)

_CATEGORY_SCHEMA_CHECKED = False
FALLBACK_CATEGORY_COLOR = '#E2E8F0'
DEFAULT_MANUAL_COLOR = '#2C6B4F'
UNCATEGORIZED_LABEL = 'Uncategorized'
CENT = Decimal('0.01')


def ensure_category_schema(force=False):
    global _CATEGORY_SCHEMA_CHECKED
    if _CATEGORY_SCHEMA_CHECKED and not force:
        return
    try:
        bind = db.engine
        inspector = inspect(bind)
        tables = inspector.get_table_names()
        if 'custom_categories' not in tables or 'category_rules' not in tables:
            current_app.logger.warning('Category schema missing required tables. Please run database migrations.')
            _CATEGORY_SCHEMA_CHECKED = True
            return

        if 'transaction_category_override' not in tables:
            stmt = text(
                """
                CREATE TABLE transaction_category_override (
                    transaction_id INTEGER NOT NULL PRIMARY KEY,
                    custom_category_id INTEGER NULL,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    CONSTRAINT fk_transaction_override_transaction
                        FOREIGN KEY (transaction_id) REFERENCES transactions (id)
                        ON DELETE CASCADE,
                    CONSTRAINT fk_transaction_override_category
                        FOREIGN KEY (custom_category_id) REFERENCES custom_categories (id)
                        ON DELETE SET NULL
                )
                """
            )
            with bind.connect() as conn:
                conn.execute(stmt)

        _CATEGORY_SCHEMA_CHECKED = True
    except Exception as exc:
        current_app.logger.warning('Unable to ensure categories schema: %s', exc)
        _CATEGORY_SCHEMA_CHECKED = False


def _with_schema_retry(handler):
    try:
        return handler()
    except OperationalError as exc:
        error_args = getattr(exc.orig, 'args', None)
        error_code = error_args[0] if error_args else None
        if error_code in (1054, 1412, 1146):
            current_app.logger.warning('Detected schema mismatch (%s). Retrying after ensuring schema.', error_code)
            db.session.rollback()
            ensure_category_schema(force=True)
            return handler()
        current_app.logger.error(f"An unhandled operational error occurred: {exc}", exc_info=True)
        return jsonify({"error": "A database error occurred. Please try again later."}), 500
    except Exception as e:
        current_app.logger.error(f"An unhandled exception occurred: {e}", exc_info=True)
        return jsonify({"error": "An unexpected error occurred. Please try again later."}), 500

_FIELD_MAP = {
    'description': 'description',
    'merchant': 'merchant',
    'category': 'category'
}

_TYPE_ALIASES = {
    'credit': 'credit',
    'deposit': 'credit',
    'income': 'credit',
    'incoming': 'credit',
    'debit': 'debit',
    'withdrawal': 'debit',
    'expense': 'debit',
    'outgoing': 'debit'
}


def _extract_category_list(raw_value):
    if isinstance(raw_value, list):
        return [str(item) for item in raw_value if item]
    if raw_value is None:
        return []
    if isinstance(raw_value, str):
        try:
            parsed = json.loads(raw_value)
            if isinstance(parsed, list):
                return [str(item) for item in parsed if item]
        except (TypeError, ValueError):
            return [raw_value] if raw_value else []
        return [raw_value] if raw_value else []
    return [str(raw_value)]


def _derive_fallback_label_from_category(raw_value):
    categories = _extract_category_list(raw_value)
    if categories:
        return ' / '.join(categories)
    return UNCATEGORIZED_LABEL


def _parse_request_date(raw_value):
    if not raw_value:
        return None
    value = str(raw_value).strip()
    if not value:
        return None
    # Normalize common separators to hyphen for easier matching.
    normalized = value.replace('/', '-').replace('.', '-')
    # Attempt ISO parsing first.
    try:
        return date.fromisoformat(normalized)
    except (ValueError, AttributeError):
        pass
    # Try a small set of alternative day/month orderings to accommodate locale-formatted values.
    for fmt in ('%Y-%m-%d', '%m-%d-%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(normalized, fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def _parse_request_decimal(raw_value):
    if raw_value is None:
        return None
    if isinstance(raw_value, (int, float, Decimal)):
        try:
            return Decimal(str(raw_value))
        except (InvalidOperation, TypeError, ValueError):
            return None
    value = str(raw_value).strip()
    if not value:
        return None
    # Remove grouping separators and normalize unicode minus signs.
    normalized = (
        value.replace(',', '')
        .replace(' ', '')
        .replace('\u2212', '-')
        .replace('\u2013', '-')
        .replace('\u2014', '-')
    )
    try:
        return Decimal(normalized)
    except (InvalidOperation, TypeError, ValueError):
        return None


def _parse_transaction_filters(args):
    account_ids_param = args.get('account_ids', '') or ''
    account_ids = []
    for part in account_ids_param.split(','):
        value = part.strip()
        if value.isdigit():
            account_ids.append(int(value))

    try:
        page = int(args.get('page', 1))
    except (TypeError, ValueError):
        page = 1
    page = max(page, 1)

    try:
        page_size = int(args.get('page_size', 200))
    except (TypeError, ValueError):
        page_size = 200
    page_size = max(1, min(page_size, 500))

    sort_key = args.get('sort_key', 'date')
    sort_desc_value = args.get('sort_desc', 'true')
    sort_desc = str(sort_desc_value).lower() in ('true', '1', 'yes', 'y', 'desc')
    search_term = (args.get('search') or '').strip()

    start_date = _parse_request_date((args.get('start_date') or '').strip())
    end_date = _parse_request_date((args.get('end_date') or '').strip())
    if start_date and end_date and start_date > end_date:
        start_date, end_date = end_date, start_date

    min_amount = _parse_request_decimal((args.get('min_amount') or '').strip())
    max_amount = _parse_request_decimal((args.get('max_amount') or '').strip())

    custom_category_param = (args.get('custom_category_id') or '').strip()

    return {
        'account_ids': account_ids,
        'page': page,
        'page_size': page_size,
        'sort_key': sort_key,
        'sort_desc': sort_desc,
        'search_term': search_term,
        'start_date': start_date,
        'end_date': end_date,
        'min_amount': min_amount,
        'max_amount': max_amount,
        'custom_category_param': custom_category_param
    }


def _build_transactions_query(user_id, filters):
    override_alias = aliased(TransactionCategoryOverride)
    query = Transaction.query.options(
        joinedload(Transaction.account),
        joinedload(Transaction.credential),
        joinedload(Transaction.custom_category)
    ).join(Account).join(Credential).outerjoin(
        override_alias, override_alias.transaction_id == Transaction.id
    ).filter(
        Transaction.user_id == user_id,
        Transaction.is_removed.is_(False),
        or_(
            Transaction.has_split_children.is_(False),
            Transaction.is_split_child.is_(True)
        )
    )

    account_ids = filters.get('account_ids') or []
    if account_ids:
        query = query.filter(Transaction.account_id.in_(account_ids))

    search_term = filters.get('search_term')
    if search_term:
        token = f"%{search_term}%"
        query = query.filter(or_(
            Transaction.name.ilike(token),
            Transaction.merchant_name.ilike(token),
            Transaction.category.ilike(token),
            func.coalesce(Transaction.payment_channel, '').ilike(token),
            Account.name.ilike(token),
            func.coalesce(Account.mask, '').ilike(token),
            Credential.institution_name.ilike(token),
            Transaction.custom_category.has(CustomCategory.name.ilike(token))
        ))

    start_date = filters.get('start_date')
    if start_date:
        query = query.filter(Transaction.date >= start_date)
    end_date = filters.get('end_date')
    if end_date:
        query = query.filter(Transaction.date <= end_date)

    min_amount = filters.get('min_amount')
    if min_amount is not None:
        query = query.filter(Transaction.amount >= min_amount)
    max_amount = filters.get('max_amount')
    if max_amount is not None:
        query = query.filter(Transaction.amount <= max_amount)

    custom_category_param = filters.get('custom_category_param', '')
    if custom_category_param:
        normalized = custom_category_param.lower()
        if normalized in ('none', 'null', 'uncategorized'):
            query = query.filter(
                Transaction.custom_category_id.is_(None),
                override_alias.custom_category_id.is_(None)
            )
        else:
            try:
                custom_category_id = int(custom_category_param)
            except (TypeError, ValueError):
                custom_category_id = None
            if custom_category_id is not None:
                query = query.filter(or_(
                    Transaction.custom_category_id == custom_category_id,
                    override_alias.custom_category_id == custom_category_id
                ))

    return query


def _normalize_category_name(value):
    if value is None:
        return ''
    return str(value).strip()


def _validate_category_name(raw_name, *, plaid_candidates=None, min_length=3):
    trimmed = _normalize_category_name(raw_name)
    if len(trimmed) < min_length:
        raise ValueError(f'Category name must be at least {min_length} characters.')
    if plaid_candidates:
        lowered = trimmed.lower()
        for candidate in plaid_candidates:
            if not candidate:
                continue
            candidate_trimmed = _normalize_category_name(candidate)
            if not candidate_trimmed:
                continue
            if lowered == candidate_trimmed.lower():
                raise ValueError('Category name cannot match an original Automatic category.')
    return trimmed


def _find_custom_category(user_id, label):
    normalized = _normalize_category_name(label)
    if not normalized:
        return None
    return CustomCategory.query.filter(
        CustomCategory.user_id == user_id,
        func.lower(CustomCategory.name) == normalized.lower()
    ).first()


def _get_or_create_custom_category(user_id, label, *, color=None, plaid_candidates=None):
    normalized_input = _normalize_category_name(label)
    category = _find_custom_category(user_id, normalized_input)
    if category:
        if color is not None:
            try:
                normalized_color = _normalize_color(color)
            except ValueError:
                normalized_color = None
            if normalized_color and category.color != normalized_color:
                category.color = normalized_color
        return category

    cleaned = _validate_category_name(normalized_input, plaid_candidates=plaid_candidates)
    normalized_color = None
    if color is not None:
        try:
            normalized_color = _normalize_color(color)
        except ValueError:
            normalized_color = None

    category = CustomCategory(
        user_id=user_id,
        name=cleaned,
        color=normalized_color or DEFAULT_MANUAL_COLOR
    )
    db.session.add(category)
    db.session.flush()
    return category


def _load_overrides(transaction_ids):
    if not transaction_ids:
        return {}
    try:
        overrides = TransactionCategoryOverride.query.options(
            joinedload(TransactionCategoryOverride.custom_category)
        ).filter(
            TransactionCategoryOverride.transaction_id.in_(transaction_ids)
        ).all()
    except Exception:
        ensure_category_schema(force=True)
        overrides = TransactionCategoryOverride.query.options(
            joinedload(TransactionCategoryOverride.custom_category)
        ).filter(
            TransactionCategoryOverride.transaction_id.in_(transaction_ids)
        ).all()
    return {override.transaction_id: override for override in overrides}


def _serialize_transaction(txn, override_map):
    category_list = _extract_category_list(txn.category)
    fallback_label = ' / '.join(category_list) if category_list else UNCATEGORIZED_LABEL

    override = override_map.get(txn.id)
    if override and override.custom_category:
        label = override.custom_category.name or fallback_label
        color = override.custom_category.color or DEFAULT_MANUAL_COLOR
        source = 'manual'
    elif txn.custom_category is not None:
        label = txn.custom_category.name or fallback_label
        color = txn.custom_category.color or DEFAULT_MANUAL_COLOR
        source = 'rule'
    else:
        label = fallback_label
        color = FALLBACK_CATEGORY_COLOR
        source = 'fallback'

    return {
        'id': txn.id,
        'plaid_transaction_id': txn.plaid_transaction_id,
        'date': txn.date.isoformat() if txn.date else None,
        'name': txn.name,
        'amount': float(txn.amount) if txn.amount is not None else 0.0,
        'iso_currency_code': txn.iso_currency_code,
        'category': category_list,
        'merchant_name': txn.merchant_name,
        'payment_channel': txn.payment_channel,
        'pending': bool(txn.pending),
        'account_id': txn.account_id,
        'account_name': txn.account.name if txn.account else None,
        'account_mask': txn.account.mask if txn.account else None,
        'account_type': txn.account.type if txn.account else None,
        'account_subtype': txn.account.subtype if txn.account else None,
        'institution_name': txn.credential.institution_name if txn.credential else None,
        'custom_category_id': txn.custom_category_id,
        'custom_category': label,
        'custom_category_color': color,
        'custom_category_source': source,
        'custom_category_is_fallback': source == 'fallback',
        'parent_transaction_id': txn.parent_transaction_id,
        'is_split_child': bool(getattr(txn, 'is_split_child', False)),
        'has_split_children': bool(getattr(txn, 'has_split_children', False)),
        'split_children_count': txn.split_children.count() if getattr(txn, 'has_split_children', False) else 0,
        'is_new': bool(txn.is_new),
        'seen_by_user': bool(txn.seen_by_user)
    }

_COLOR_RE = re.compile(r'^#([0-9a-fA-F]{6})$')


def _wildcard_to_regex(pattern):
    escaped = re.escape(pattern)
    escaped = escaped.replace(r'\*', '.*')
    escaped = escaped.replace(r'\?', '.')
    escaped = escaped.replace(r'\#', r'\d')
    return re.compile(escaped, re.IGNORECASE)


def _resolve_rule_field(field_value):
    if not field_value:
        return 'description'
    field_value = field_value.lower()
    return _FIELD_MAP.get(field_value, 'description')


def _resolve_rule_type(type_value):
    if not type_value:
        return None
    normalized = _TYPE_ALIASES.get(type_value.lower(), type_value.lower())
    return normalized if normalized in {'credit', 'debit'} else None


def _extract_field_value(transaction, field_name):
    if field_name == 'merchant':
        return transaction.merchant_name or ''
    if field_name == 'category':
        raw = transaction.category or ''
        if isinstance(raw, list):
            return ' '.join([str(item) for item in raw if item])
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, list):
                return ' '.join([str(item) for item in parsed if item])
            if parsed:
                return str(parsed)
        except (TypeError, ValueError):
            pass
        return str(raw) if raw else ''
    return transaction.name or ''


def _determine_transaction_flow(amount):
    if amount is None:
        return None
    try:
        value = Decimal(str(amount))
    except (InvalidOperation, TypeError, ValueError):
        return None
    if value < 0:
        return 'debit'
    if value > 0:
        return 'credit'
    return 'neutral'


def _normalize_color(value):
    if not value:
        return '#2C6B4F'
    candidate = value.strip()
    if not candidate.startswith('#'):
        candidate = f'#{candidate}'
    candidate = candidate.upper()
    if not _COLOR_RE.match(candidate):
        raise ValueError('Invalid color value')
    return candidate


def _compile_category_rules(user_id):
    ensure_category_schema()
    rules = CategoryRule.query.options(
        joinedload(CategoryRule.category)
    ).filter_by(user_id=user_id).order_by(CategoryRule.created_at.asc(), CategoryRule.id.asc()).all()
    compiled = []
    for rule in rules:
        category = rule.category
        pattern = (rule.text_to_match or '').strip()
        if not pattern or category is None:
            continue
        try:
            regex = _wildcard_to_regex(pattern)
        except re.error:
            continue
        label = (category.name or '').strip()
        if not label:
            continue
        compiled.append({
            'rule_id': rule.id,
            'category_id': category.id,
            'regex': regex,
            'field': _resolve_rule_field(rule.field_to_match),
            'type': _resolve_rule_type(rule.transaction_type),
            'amount_min': Decimal(str(rule.amount_min)) if rule.amount_min is not None else None,
            'amount_max': Decimal(str(rule.amount_max)) if rule.amount_max is not None else None,
            'label': label,
            'color': category.color or DEFAULT_MANUAL_COLOR
        })
    return compiled


def apply_rules_to_transactions(user_id, transaction_ids=None, include_removed=False):
    """
    Apply category rules to the user's transactions. If transaction_ids is provided,
    only those transactions are evaluated.
    """
    valid_rules = _compile_category_rules(user_id)
    rule_count = len(valid_rules)

    transactions_query = Transaction.query.filter(Transaction.user_id == user_id)
    if transaction_ids:
        transactions_query = transactions_query.filter(Transaction.id.in_(transaction_ids))
    if not include_removed:
        transactions_query = transactions_query.filter(Transaction.is_removed.is_(False))
    transactions = transactions_query.all()
    override_map = _load_overrides([txn.id for txn in transactions])

    updated = 0
    matched = 0

    for txn in transactions:
        if txn.id in override_map:
            continue

        txn_amount = Decimal(str(txn.amount)) if txn.amount is not None else None
        abs_amount = txn_amount.copy_abs() if txn_amount is not None else None
        txn_flow = _determine_transaction_flow(txn_amount)
        chosen_rule = None

        for compiled in valid_rules:
            if compiled['type'] and compiled['type'] != txn_flow:
                continue

            if compiled['amount_min'] is not None and (abs_amount is None or abs_amount < compiled['amount_min']):
                continue
            if compiled['amount_max'] is not None and (abs_amount is None or abs_amount > compiled['amount_max']):
                continue

            field_value = _extract_field_value(txn, compiled['field'])
            if not field_value:
                continue

            if compiled['regex'].search(field_value):
                chosen_rule = compiled

        if chosen_rule is not None:
            category_id = chosen_rule['category_id']
            if txn.custom_category_id != category_id:
                txn.custom_category_id = category_id
                updated += 1
            matched += 1
        else:
            if txn.custom_category_id is not None:
                txn.custom_category_id = None
                updated += 1

    return {
        'rules_processed': rule_count,
        'matched_transactions': matched,
        'transactions_updated': updated
    }


def apply_category_rules(user_id, include_removed=False):
    """
    Apply category rules to all of the user's transactions.
    """
    return apply_rules_to_transactions(user_id, transaction_ids=None, include_removed=include_removed)


def _serialize_custom_category(category, extras=None):
    if not category:
        return None
    data = {
        'id': category.id,
        'name': category.name,
        'color': category.color,
        'created_at': category.created_at.isoformat() if category.created_at else None,
        'updated_at': category.updated_at.isoformat() if category.updated_at else None,
    }
    if extras:
        data.update(extras)
    return data


def _serialize_category(rule):
    category = getattr(rule, 'category', None)
    return {
        'id': rule.id,
        'text_to_match': rule.text_to_match,
        'field_to_match': rule.field_to_match,
        'transaction_type': rule.transaction_type,
        'amount_min': str(rule.amount_min) if rule.amount_min is not None else None,
        'amount_max': str(rule.amount_max) if rule.amount_max is not None else None,
        'category_id': category.id if category else None,
        'category': _serialize_custom_category(category),
        'created_at': rule.created_at.isoformat() if rule.created_at else None,
        'updated_at': rule.updated_at.isoformat() if rule.updated_at else None
    }


def _parse_decimal_value(value):
    if value is None or value == '':
        return None
    try:
        decimal_value = Decimal(str(value))
        return decimal_value.copy_abs()
    except (InvalidOperation, TypeError, ValueError):
        raise ValueError('Invalid decimal value')


def _extract_label(payload):
    label = payload.get('label')
    if label:
        return label
    return payload.get('custom_category')


def _collect_category_labels(user_id):
    ensure_category_schema()
    labels = (
        db.session.query(CustomCategory.name)
        .filter(CustomCategory.user_id == user_id, CustomCategory.name.isnot(None))
        .distinct()
        .all()
    )
    return sorted(
        label for (label,) in labels if label
    )


def _collect_custom_category_stats(user_id, category_ids):
    if not category_ids:
        return {}, {}, {}

    rule_counts = dict(
        db.session.query(CategoryRule.category_id, func.count(CategoryRule.id))
        .filter(
            CategoryRule.user_id == user_id,
            CategoryRule.category_id.in_(category_ids)
        )
        .group_by(CategoryRule.category_id)
        .all()
    )

    transaction_counts = dict(
        db.session.query(Transaction.custom_category_id, func.count(Transaction.id))
        .filter(
            Transaction.user_id == user_id,
            Transaction.custom_category_id.in_(category_ids)
        )
        .group_by(Transaction.custom_category_id)
        .all()
    )

    override_counts = dict(
        db.session.query(TransactionCategoryOverride.custom_category_id, func.count(TransactionCategoryOverride.transaction_id))
        .join(Transaction, Transaction.id == TransactionCategoryOverride.transaction_id)
        .filter(
            Transaction.user_id == user_id,
            TransactionCategoryOverride.custom_category_id.in_(category_ids)
        )
        .group_by(TransactionCategoryOverride.custom_category_id)
        .all()
    )

    return rule_counts, transaction_counts, override_counts


def _resolve_transaction_category_label(txn, override_map, allow_fallback=True):
    label = None
    if override_map:
        override = override_map.get(txn.id)
        if override and override.custom_category:
            label = override.custom_category.name
    if not label and getattr(txn, 'custom_category', None):
        label = txn.custom_category.name
    if not label and allow_fallback:
        fallback = _extract_category_list(txn.category)
        if fallback:
            label = ' / '.join(fallback)
    if label:
        return label.strip()
    return None


_BUDGET_FREQUENCY_MAP = {
    'weekly': 'Weekly',
    'biweekly': 'Biweekly',
    'semi-monthly': 'Semi-Monthly',
    'monthly': 'Monthly',
    'quarterly': 'Quarterly',
    'yearly': 'Yearly'
}


def _normalize_budget_frequency(value):
    if not value:
        return 'monthly'
    normalized = str(value).strip().lower()
    if normalized in _BUDGET_FREQUENCY_MAP:
        return normalized
    raise ValueError('Invalid frequency value.')


def _normalize_budget_amount(value):
    if value is None or value == '':
        raise ValueError('Amount is required.')
    try:
        decimal_value = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        raise ValueError('Invalid amount value.')
    decimal_value = decimal_value.copy_abs().quantize(CENT, rounding=ROUND_HALF_UP)
    if decimal_value <= Decimal('0.00'):
        raise ValueError('Amount must be greater than zero.')
    return decimal_value


def _month_start_offset(reference_date, offset):
    year = reference_date.year
    month = reference_date.month - offset
    while month <= 0:
        month += 12
        year -= 1
    return date(year, month, 1)


def _determine_budget_classification(avg_value, current_value):
    if current_value > 0 or avg_value > 0:
        return 'income'
    if current_value < 0 or avg_value < 0:
        return 'expense'
    return 'expense'


def _calculate_budget_metrics(user_id, category_labels):
    normalized_labels = {
        (label or '').strip().lower(): label
        for label in category_labels if label and label.strip()
    }
    if not normalized_labels:
        return {}

    today = date.today()
    month_starts = [_month_start_offset(today, offset) for offset in range(0, 6)]
    earliest_start = month_starts[-1]
    month_keys = [start.strftime('%Y-%m') for start in month_starts]
    current_month_key = month_keys[0]

    transactions = Transaction.query.options(
        joinedload(Transaction.custom_category)
    ).filter(
        Transaction.user_id == user_id,
        Transaction.is_removed.is_(False),
        Transaction.date >= earliest_start
    ).all()

    override_map = _load_overrides([txn.id for txn in transactions])
    monthly_totals = defaultdict(lambda: defaultdict(lambda: Decimal('0.00')))

    for txn in transactions:
        if txn.has_split_children and not txn.is_split_child:
            continue
        label = _resolve_transaction_category_label(txn, override_map, allow_fallback=False)
        if not label:
            continue
        label_key = label.lower()
        if label_key not in normalized_labels:
            continue
        if not txn.date:
            continue
        month_key = txn.date.strftime('%Y-%m')
        if month_key not in month_keys:
            continue
        amount = Decimal(txn.amount or 0).quantize(CENT, rounding=ROUND_HALF_UP)
        monthly_totals[label_key][month_key] += amount

    metrics = {}
    for label_key in normalized_labels:
        totals_by_month = monthly_totals.get(label_key, {})
        month_values = [totals_by_month.get(key, Decimal('0.00')) for key in month_keys]
        non_zero_values = [value for value in month_values if value != Decimal('0.00')]
        if non_zero_values:
            average_value = (sum(non_zero_values, Decimal('0.00')) / Decimal(len(non_zero_values))).quantize(CENT, rounding=ROUND_HALF_UP)
        else:
            average_value = Decimal('0.00')
        current_value = totals_by_month.get(current_month_key, Decimal('0.00')).quantize(CENT, rounding=ROUND_HALF_UP)
        classification = _determine_budget_classification(average_value, current_value)
        metrics[label_key] = {
            'six_month_average': average_value,
            'current_month_total': current_value,
            'classification': classification
        }
    return metrics


def _serialize_budget(budget, metrics_map):
    label_key = (budget.category_label or '').strip().lower()
    metrics = metrics_map.get(label_key, {})
    avg_value = metrics.get('six_month_average', Decimal('0.00'))
    current_value = metrics.get('current_month_total', Decimal('0.00'))
    classification = metrics.get('classification', _determine_budget_classification(Decimal('0.00'), Decimal('0.00')))

    return {
        'id': budget.id,
        'category_label': budget.category_label,
        'frequency': budget.frequency,
        'amount': str(Decimal(budget.amount or 0).quantize(CENT, rounding=ROUND_HALF_UP)),
        'six_month_average': float(avg_value),
        'current_month_total': float(current_value),
        'classification': classification,
        'created_at': budget.created_at.isoformat() if budget.created_at else None,
        'updated_at': budget.updated_at.isoformat() if budget.updated_at else None
    }


def _build_budget_summary(budgets, metrics_map):
    income_total = Decimal('0.00')
    expense_total = Decimal('0.00')
    for budget in budgets:
        amount = Decimal(budget.amount or 0).quantize(CENT, rounding=ROUND_HALF_UP)
        label_key = (budget.category_label or '').strip().lower()
        metrics = metrics_map.get(label_key, {})
        classification = metrics.get('classification', 'expense')
        if classification == 'income':
            income_total += amount
        else:
            expense_total += amount
    net_total = income_total - expense_total
    return {
        'income_total': float(income_total),
        'expense_total': float(expense_total),
        'net_total': float(net_total)
    }


_ACCOUNT_GROUP_LABELS = {
    'banking': 'Banking',
    'credit': 'Credit',
    'investment': 'Investment',
    'other': 'Other'
}


_INVESTMENT_TYPES = {
    'investment', 'brokerage', '401k', 'ira', 'retirement', 'education'
}


def _classify_account_group(account):
    type_value = (account.type or '').lower()
    subtype_value = (account.subtype or '').lower()
    if type_value in {'credit', 'loan'}:
        return 'credit'
    if type_value in {'depository', 'checking', 'savings'}:
        return 'banking'
    if type_value in _INVESTMENT_TYPES or subtype_value in _INVESTMENT_TYPES:
        return 'investment'
    return 'other'


def _format_decimal(value):
    if isinstance(value, Decimal):
        return float(value.quantize(CENT, rounding=ROUND_HALF_UP))
    return float(Decimal(value or 0).quantize(CENT, rounding=ROUND_HALF_UP))


def _collect_balances_summary(user_id):
    transaction_totals = db.session.query(
        Transaction.account_id.label('account_id'),
        func.coalesce(func.sum(Transaction.amount), Decimal('0.00')).label('balance')
    ).filter(
        Transaction.user_id == user_id,
        Transaction.is_removed.is_(False)
    ).group_by(Transaction.account_id).subquery()

    accounts = (
        db.session.query(
            Account,
            Credential.institution_name,
            func.coalesce(transaction_totals.c.balance, Decimal('0.00')).label('balance')
        )
        .join(Credential, Account.credential_id == Credential.id)
        .outerjoin(transaction_totals, transaction_totals.c.account_id == Account.id)
        .filter(Credential.user_id == user_id, Account.is_enabled.is_(True))
        .all()
    )

    groups = {}
    grand_total = Decimal('0.00')
    processed_accounts = set()

    for account, institution_name, balance in accounts:
        account_key = (institution_name, account.name, account.mask)
        if account_key in processed_accounts:
            continue
        processed_accounts.add(account_key)
        
        group_key = _classify_account_group(account)
        if group_key not in groups:
            groups[group_key] = {
                'key': group_key,
                'label': _ACCOUNT_GROUP_LABELS.get(group_key, group_key.title()),
                'total': Decimal('0.00'),
                'institutions': {}
            }

        account_balance = Decimal(balance or 0)
        if (account.type or '').lower() in {'credit', 'loan'}:
            account_balance = -account_balance
        account_balance = account_balance.quantize(CENT, rounding=ROUND_HALF_UP)

        institution_key = f"{group_key}:{institution_name or 'Unknown Institution'}"
        institution_entry = groups[group_key]['institutions'].setdefault(institution_key, {
            'id': institution_key,
            'name': institution_name or 'Unknown Institution',
            'total': Decimal('0.00'),
            'accounts': []
        })

        account_entry = {
            'id': account.id,
            'name': account.name or 'Account',
            'mask': account.mask,
            'balance': float(account_balance)
        }
        institution_entry['accounts'].append(account_entry)
        institution_entry['total'] += account_balance
        groups[group_key]['total'] += account_balance
        grand_total += account_balance

    group_list = []
    for group_key, group in groups.items():
        institutions = []
        for institution in group['institutions'].values():
            institution['accounts'].sort(key=lambda acc: acc['name'])
            institution['total'] = float(institution['total'].quantize(CENT, rounding=ROUND_HALF_UP))
            institutions.append(institution)
        institutions.sort(key=lambda inst: inst['name'])
        group_list.append({
            'key': group['key'],
            'label': group['label'],
            'total': float(group['total'].quantize(CENT, rounding=ROUND_HALF_UP)),
            'institutions': institutions
        })

    group_list.sort(key=lambda item: item['label'])

    return {
        'groups': group_list,
        'grand_total': float(grand_total.quantize(CENT, rounding=ROUND_HALF_UP))
    }


def _calculate_spending_metrics(user_id, category_labels):
    """
    Calculates 6-month average and current month total for spending categories,
    only considering expenses (negative amounts). Used specifically for the spending report.
    """
    normalized_labels = {
        (label or '').strip().lower(): label
        for label in category_labels if label and label.strip()
    }
    if not normalized_labels:
        return {}

    today = date.today()
    month_starts = [_month_start_offset(today, offset) for offset in range(0, 6)]
    earliest_start = month_starts[-1]
    month_keys = [start.strftime('%Y-%m') for start in month_starts]
    current_month_key = month_keys[0]

    transactions = Transaction.query.options(
        joinedload(Transaction.custom_category)
    ).filter(
        Transaction.user_id == user_id,
        Transaction.is_removed.is_(False),
        Transaction.date >= earliest_start,
        Transaction.amount < 0  # Only consider expenses
    ).all()

    override_map = _load_overrides([txn.id for txn in transactions])
    monthly_spending_totals = defaultdict(lambda: defaultdict(lambda: Decimal('0.00')))

    for txn in transactions:
        if txn.has_split_children and not txn.is_split_child:
            continue
        label = _resolve_transaction_category_label(txn, override_map, allow_fallback=False)
        if not label:
            continue
        label_key = label.lower()
        if label_key not in normalized_labels:
            continue
        if not txn.date:
            continue
        month_key = txn.date.strftime('%Y-%m')
        if month_key not in month_keys:
            continue
        amount = Decimal(txn.amount or 0).quantize(CENT, rounding=ROUND_HALF_UP)
        monthly_spending_totals[label_key][month_key] += abs(amount)  # Sum absolute values of expenses

    metrics = {}
    for label_key in normalized_labels:
        totals_by_month = monthly_spending_totals.get(label_key, {})
        month_values = [totals_by_month.get(key, Decimal('0.00')) for key in month_keys]
        non_zero_values = [value for value in month_values if value != Decimal('0.00')]
        if non_zero_values:
            average_value = (sum(non_zero_values, Decimal('0.00')) / Decimal(len(non_zero_values))).quantize(CENT, rounding=ROUND_HALF_UP)
        else:
            average_value = Decimal('0.00')
        current_value = totals_by_month.get(current_month_key, Decimal('0.00')).quantize(CENT, rounding=ROUND_HALF_UP)
        
        # For spending report, classification is always 'expense' if there's spending
        classification = 'expense' if current_value > 0 or average_value > 0 else 'expense'

        metrics[label_key] = {
            'six_month_average': average_value,
            'current_month_total': current_value,
            'classification': classification
        }
    return metrics


def _collect_spending_summary(user_id):
    today = date.today()
    current_year = today.year
    current_month = today.month
    start_date = _month_start_offset(today, 23)

    transactions = Transaction.query.options(
        joinedload(Transaction.custom_category)
    ).filter(
        Transaction.user_id == user_id,
        Transaction.is_removed.is_(False),
        Transaction.date >= start_date
    ).order_by(Transaction.date.desc()).all()

    override_map = _load_overrides([txn.id for txn in transactions])

    # First, find all unique category labels from the transaction set
    all_category_labels = set()
    for txn in transactions:
        label = _resolve_transaction_category_label(txn, override_map, allow_fallback=False)
        if label:
            all_category_labels.add(label)

    category_entries = {}
    for label in sorted(list(all_category_labels)):
        label_key = label.lower()
        category_entries[label_key] = {
            'label': label,
            'totals': defaultdict(lambda: defaultdict(lambda: Decimal('0.00'))),
            'raw_sum': Decimal('0.00')
        }

    monthly_cashflow = defaultdict(lambda: defaultdict(lambda: Decimal('0.00')))
    uncategorized_totals = defaultdict(lambda: defaultdict(lambda: Decimal('0.00')))

    for txn in transactions:
        if txn.has_split_children and not txn.is_split_child:
            continue
        if not txn.date:
            continue
        label = _resolve_transaction_category_label(txn, override_map, allow_fallback=False)
        year = txn.date.year
        month = txn.date.month
        value = Decimal(txn.amount or 0).quantize(CENT, rounding=ROUND_HALF_UP)
        monthly_cashflow[year][month] += value
        if not label:
            uncategorized_totals[year][month] += value
            continue
        label_key = label.lower()
        info = category_entries.setdefault(label_key, {
            'label': label,
            'totals': defaultdict(lambda: defaultdict(lambda: Decimal('0.00'))),
            'raw_sum': Decimal('0.00')
        })
        if value < 0:
            spending_value = abs(value)
            info['totals'][year][month] += spending_value
        info['raw_sum'] += value

    metrics = _calculate_budget_metrics(user_id, [info['label'] for info in category_entries.values()])
    budgets = Budget.query.filter_by(user_id=user_id).all()
    budget_map = {
        (budget.category_label or '').strip().lower(): Decimal(budget.amount or 0).quantize(CENT, rounding=ROUND_HALF_UP)
        for budget in budgets
    }

    year_map = defaultdict(lambda: defaultdict(lambda: {
        'month': None,
        'label': None,
        'total': Decimal('0.00'),
        'categories': []
    }))

    for label_key, info in category_entries.items():
        metrics_entry = metrics.get(label_key, {})
        six_month_average = abs(metrics_entry.get('six_month_average', Decimal('0.00')))
        classification = metrics_entry.get('classification')
        if classification is None:
            classification = 'income' if info['raw_sum'] > 0 else 'expense'
        
        # Skip pure income categories from the spending report
        if classification == 'income' and not any(month for month in info['totals'].values() if any(val > 0 for val in month.values())):
            continue

        budget_amount = budget_map.get(label_key)

        for year, months in info['totals'].items():
            for month, value in months.items():
                month_entry = year_map[year][month]
                month_entry['month'] = month
                month_entry['label'] = calendar.month_abbr[month]
                month_entry['total'] += value
                remainder = None
                if budget_amount is not None:
                    remainder = budget_amount - value
                month_entry['categories'].append({
                    'label': info['label'],
                    'value': value,
                    'six_month_average': six_month_average,
                    'budget': budget_amount,
                    'remainder': remainder,
                    'classification': classification
                })

    # Add uncategorized totals as a synthetic category per month
    for year, months in uncategorized_totals.items():
        for month, value in months.items():
            if value == Decimal('0.00'):
                continue
            month_entry = year_map[year][month]
            month_entry['month'] = month
            month_entry['label'] = calendar.month_abbr[month]
            month_entry['total'] += value
            net_classification = 'income' if value > 0 else 'expense'
            month_entry['categories'].append({
                'label': UNCATEGORIZED_LABEL,
                'value': value,
                'six_month_average': Decimal('0.00'),
                'budget': None,
                'remainder': None,
                'classification': net_classification
            })

    for year, months in year_map.items():
        for month, entry in months.items():
            net_value = monthly_cashflow[year][month].quantize(CENT, rounding=ROUND_HALF_UP)
            entry['total'] = net_value

    years_output = []
    for year in sorted(year_map.keys(), reverse=True):
        months_output = []
        for month in sorted(year_map[year].keys(), reverse=True):
            month_entry = year_map[year][month]
            categories = month_entry['categories']
            categories.sort(key=lambda item: item['label'].lower())
            months_output.append({
                'year': year,
                'month': month,
                'label': month_entry['label'],
                'total': float(month_entry['total'].quantize(CENT, rounding=ROUND_HALF_UP)),
                'categories': [
                    {
                        'label': category['label'],
                        'value': float(category['value'].quantize(CENT, rounding=ROUND_HALF_UP)),
                        'six_month_average': float(category['six_month_average'].quantize(CENT, rounding=ROUND_HALF_UP)),
                        'budget': float(category['budget'].quantize(CENT, rounding=ROUND_HALF_UP)) if category['budget'] is not None else None,
                        'remainder': float(category['remainder'].quantize(CENT, rounding=ROUND_HALF_UP)) if category['remainder'] is not None else None,
                        'classification': category['classification']
                    }
                    for category in categories
                ]
            })
        years_output.append({
            'year': year,
            'months': months_output
        })

    return {
        'years': years_output,
        'current_year': current_year,
        'current_month': current_month,
        'generated_on': today.isoformat()
    }


def _collect_cashflow_summary(user_id):
    today = date.today()
    months = []
    month_order = []
    month_map = {}

    for offset in range(11, -1, -1):
        start = _month_start_offset(today, offset)
        key = start.strftime('%Y-%m')
        label = f"{calendar.month_abbr[start.month]} {start.year}"
        month_order.append(key)
        month_map[key] = {
            'key': key,
            'label': label,
            'income': Decimal('0.00'),
            'expense': Decimal('0.00')
        }

    start_date = _month_start_offset(today, 11)

    transactions = Transaction.query.options(
        joinedload(Transaction.custom_category)
    ).filter(
        Transaction.user_id == user_id,
        Transaction.is_removed.is_(False),
        Transaction.date >= start_date
    ).order_by(Transaction.date).all()

    override_map = _load_overrides([txn.id for txn in transactions])
    category_series = defaultdict(lambda: defaultdict(lambda: Decimal('0.00')))
    category_labels = {}

    for txn in transactions:
        if txn.has_split_children and not txn.is_split_child:
            continue
        if not txn.date:
            continue
        month_key = txn.date.strftime('%Y-%m')
        if month_key not in month_map:
            continue
        amount = Decimal(txn.amount or 0).quantize(CENT, rounding=ROUND_HALF_UP)
        if amount >= 0:
            month_map[month_key]['income'] += amount
        else:
            month_map[month_key]['expense'] += amount

        label = _resolve_transaction_category_label(txn, override_map, allow_fallback=False)
        if label:
            label_key = label.lower()
            category_series[label_key][month_key] += amount
            category_labels[label_key] = label

    months_output = []
    for key in month_order:
        entry = month_map[key]
        net_total = entry['income'] + entry['expense']
        income_val = entry['income'].quantize(CENT, rounding=ROUND_HALF_UP)
        expense_val = entry['expense'].quantize(CENT, rounding=ROUND_HALF_UP)
        net_total = income_val + expense_val

        months_output.append({
            'key': key,
            'label': entry['label'],
            'income': float(income_val),
            'expense': float(expense_val),
            'total': float(net_total),
            'series': {
                'positive': {
                    'value': float(income_val),
                    'color': '#2C6B4F'
                },
                'negative': {
                    'value': float(abs(expense_val)),
                    'color': '#A53A3D'
                }
            }
        })

    series_output = {}
    for label_key, series in category_series.items():
        series_output[label_key] = {
            month_key: float(series.get(month_key, Decimal('0.00')).quantize(CENT, rounding=ROUND_HALF_UP))
            for month_key in month_order
        }

    categories_output = [
        {'key': key, 'label': label}
        for key, label in sorted(category_labels.items(), key=lambda item: item[1].lower())
    ]

    return {
        'months': months_output,
        'categories': categories_output,
        'series': series_output
    }


def _normalize_split_amount(value):
    if value is None or value == '':
        raise ValueError('Amount is required for each split.')
    try:
        decimal_value = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        raise ValueError('Invalid amount value.')
    decimal_value = decimal_value.copy_abs().quantize(CENT, rounding=ROUND_HALF_UP)
    if decimal_value <= Decimal('0.00'):
        raise ValueError('Split amounts must be greater than zero.')
    return decimal_value


def _generate_split_transaction_id(base_identifier):
    suffix = uuid4().hex[:10]
    base = (base_identifier or 'local')[:90]
    max_base_length = max(0, 100 - len('-split-') - len(suffix))
    trimmed = base[:max_base_length] if max_base_length else ''
    if not trimmed:
        trimmed = 'local'
    return f"{trimmed}-split-{suffix}"


def _delete_existing_split_children(parent_txn):
    if not parent_txn or not parent_txn.has_split_children:
        return
    existing_children = parent_txn.split_children.all()
    for child in existing_children:
        db.session.delete(child)
    parent_txn.has_split_children = False


def _mark_split_children_removed(parent_txn):
    if not parent_txn or not parent_txn.has_split_children:
        return
    active_children = parent_txn.split_children.filter(Transaction.is_removed.is_(False)).all()
    for child in active_children:
        child.is_removed = True
        child.last_action = 'removed'
        child.updated_at = datetime.utcnow()
    parent_txn.has_split_children = False


def _rebalance_split_children(parent_txn, previous_amount, new_amount):
    if not parent_txn or not parent_txn.has_split_children:
        return

    previous_amount = Decimal(previous_amount or 0)
    new_amount = Decimal(new_amount or 0)
    previous_abs = previous_amount.copy_abs()
    new_abs = new_amount.copy_abs()

    children = parent_txn.split_children.filter(Transaction.is_removed.is_(False)).all()
    if not children:
        parent_txn.has_split_children = False
        return

    sign = -1 if new_amount < 0 or (new_amount == 0 and previous_amount < 0) else 1

    if previous_abs == 0:
        if new_abs == 0:
            for child in children:
                child.amount = Decimal('0.00')
                child.updated_at = datetime.utcnow()
                child.last_action = 'modified'
            return
        equal_share = (new_abs / len(children)).quantize(CENT, rounding=ROUND_HALF_UP)
        values = [equal_share for _ in children]
        total = sum(values, Decimal('0.00'))
        diff = (new_abs - total).quantize(CENT, rounding=ROUND_HALF_UP)
        if values:
            values[-1] = (values[-1] + diff).quantize(CENT, rounding=ROUND_HALF_UP)
    else:
        ratio = (new_abs / previous_abs) if previous_abs != 0 else Decimal('0')
        values = []
        total = Decimal('0.00')
        for child in children:
            child_abs = Decimal(child.amount or 0).copy_abs()
            scaled = (child_abs * ratio).quantize(CENT, rounding=ROUND_HALF_UP)
            values.append(scaled)
            total += scaled
        diff = (new_abs - total).quantize(CENT, rounding=ROUND_HALF_UP)
        if values:
            values[-1] = (values[-1] + diff).quantize(CENT, rounding=ROUND_HALF_UP)

    for idx, child in enumerate(children):
        adjusted = values[idx] if idx < len(values) else Decimal('0.00')
        if adjusted < 0:
            adjusted = Decimal('0.00')
        child.amount = adjusted * sign
        child.updated_at = datetime.utcnow()
        child.last_action = 'modified'


def _create_split_children(parent_txn, split_specs):
    if not parent_txn:
        raise ValueError('Parent transaction not provided.')
    if not split_specs or len(split_specs) < 2:
        raise ValueError('At least two split entries are required.')

    _delete_existing_split_children(parent_txn)
    db.session.flush()

    parent_categories = _extract_category_list(parent_txn.category)
    fallback_label = ' / '.join(parent_categories) if parent_categories else None
    plaid_candidates = list(parent_categories)
    if fallback_label:
        plaid_candidates.append(fallback_label)

    label_map = {
        _normalize_category_name(category.name).lower(): category
        for category in CustomCategory.query.filter_by(user_id=parent_txn.user_id).all()
        if category.name
    }

    parent_sign = -1 if (parent_txn.amount or Decimal('0')) < 0 else 1
    timestamp = datetime.utcnow()
    new_children = []

    for spec in split_specs:
        description = spec['description']
        category_label = spec['category']
        amount_abs = spec['amount']
        child_amount = (amount_abs * parent_sign).quantize(CENT, rounding=ROUND_HALF_UP)
        plaid_id = _generate_split_transaction_id(parent_txn.plaid_transaction_id)

        try:
            cleaned_label = _validate_category_name(category_label, plaid_candidates=plaid_candidates)
        except ValueError as exc:
            raise ValueError(str(exc))

        child = Transaction(
            plaid_transaction_id=plaid_id,
            user_id=parent_txn.user_id,
            credential_id=parent_txn.credential_id,
            account_id=parent_txn.account_id,
            name=description,
            amount=child_amount,
            iso_currency_code=parent_txn.iso_currency_code,
            category=parent_txn.category,
            merchant_name=parent_txn.merchant_name,
            payment_channel=parent_txn.payment_channel,
            date=parent_txn.date,
            pending=parent_txn.pending,
            is_removed=False,
            last_action='split',
            created_at=timestamp,
            updated_at=timestamp,
            parent_transaction_id=parent_txn.id,
            is_split_child=True,
            has_split_children=False
        )

        normalized_label = cleaned_label.lower()
        matching_category = label_map.get(normalized_label)
        if not matching_category:
            matching_category = _get_or_create_custom_category(
                parent_txn.user_id,
                cleaned_label,
                color=DEFAULT_MANUAL_COLOR,
                plaid_candidates=plaid_candidates
            )
            label_map[matching_category.name.strip().lower()] = matching_category

        child.custom_category_id = matching_category.id

        db.session.add(child)
        new_children.append(child)

    db.session.flush()

    parent_txn.has_split_children = True
    parent_txn.is_split_child = False
    parent_txn.updated_at = datetime.utcnow()
    parent_txn.last_action = 'split'
    parent_txn.custom_category_id = None

    parent_override = TransactionCategoryOverride.query.filter_by(transaction_id=parent_txn.id).first()
    if parent_override:
        db.session.delete(parent_override)

    return new_children

@core.route('/create_link_token', methods=['POST'])
@login_required
def create_link_token():
    data = request.json or {}
    access_token = data.get('access_token')
    is_refresh = data.get('is_refresh', False)
    
    link_token_request = {
        'user': {
            'client_user_id': str(current_user.id),
        },
        'client_name': "ExMint",
        'products': ["transactions"],
        'country_codes': ['CA','US'],
        'language': 'en',
        'redirect_uri': 'https://automatos.ca',
        'webhook': current_app.config['PLAID_WEBHOOK_URL'],
        "update": {
        "account_selection_enabled": True
        },
    }

    link_token_request["update"] = {"account_selection_enabled": True}
        
    if access_token:
        link_token_request['access_token'] = access_token

    try:
        response = current_app.plaid_client.link_token_create(link_token_request)
        return jsonify(response.to_dict())
    except plaid.ApiException as e:
        return jsonify({'error': str(e)})

@core.route('/handle_token_and_accounts', methods=['POST'])
@login_required
def handle_token_and_accounts():
    data = request.json
    credential_id = data.get('credential_id')
    public_token = data.get('public_token', None)
    is_refresh = data.get('is_refresh', False)

    try:
        # Fetch the webhook URL from the environment (via Config)
        webhook_url = current_app.config['PLAID_WEBHOOK_URL']
        if not webhook_url:
            return jsonify({'error': 'Webhook URL is not configured in the environment.'}), 500

        if is_refresh:
            print(f"Trying to refresh connection...")
            credential = Credential.query.get(credential_id)
            if not credential:
                return jsonify({'error': 'Credential not found'}), 404
            
            access_token = credential.access_token
            print(f"** Debug: Refreshing with access_token: {access_token}")

            # Update the webhook for the existing item
            webhook_update_request = {
                'client_id': current_app.config['PLAID_CLIENT_ID'],
                'secret': current_app.config['PLAID_SECRET'],
                'access_token': access_token,
                'webhook': webhook_url
            }
            webhook_update_response = current_app.plaid_client.item_webhook_update(webhook_update_request)
            webhook_update_data = webhook_update_response.to_dict()
            print(f"** Debug: Webhook updated for refresh. Response: {webhook_update_data}")

            # Extract and update the item_id in the Credential table
            item_id = webhook_update_data['item']['item_id']
            print(f"Item ID during refresh: {item_id}")
            credential.item_id = item_id
            db.session.commit()
            print(f"Connection refreshed...")

        else:
            print(f"Trying to add new connection...")
            public_token = data['public_token']
            institution_name = data.get('institution_name', 'Unknown')
            exchange_request = ItemPublicTokenExchangeRequest(public_token=public_token)
            exchange_response = current_app.plaid_client.item_public_token_exchange(exchange_request)
            access_token = exchange_response['access_token']
            item_id = exchange_response['item_id']  # Extract item_id
            print(f"Item ID during creation: {item_id}")

            credential = Credential(
                user_id=current_user.id,
                access_token=access_token,
                institution_name=institution_name,
                item_id=item_id,  # Store the item_id
                requires_update=False
            )
            db.session.add(credential)
            db.session.commit()

            credential_id = credential.id

        # Fetch accounts using the access token
        accounts_request = AccountsGetRequest(access_token=access_token)
        accounts_response = current_app.plaid_client.accounts_get(accounts_request)

        accounts_data = accounts_response.to_dict().get('accounts', [])
        for account in accounts_data:
            plaid_account_id = account.get('account_id')
            name = account.get('name')
            type_ = account.get('type')
            subtype = account.get('subtype')
            mask = account.get('mask')
            is_enabled = account.get('is_enabled', True)  # Default to True if not present

            existing_account = Account.query.filter_by(plaid_account_id=plaid_account_id).first()
            if existing_account:
                existing_account.name = name
                existing_account.type = type_
                existing_account.subtype = subtype
                existing_account.mask = mask
                existing_account.is_enabled = is_enabled
            else:
                new_account = Account(
                    status='Active', 
                    credential_id=credential_id,
                    plaid_account_id=plaid_account_id,
                    name=name,
                    type=type_,
                    subtype=subtype,
                    mask=mask,
                    is_enabled=is_enabled
                )
                db.session.add(new_account)
        db.session.commit()

        filtered_response = {
            'item': accounts_response['item'],
            'request_id': accounts_response['request_id']
        }

        db.session.refresh(credential)

        # Clear any update flags and run an initial sync so that Plaid webhooks can begin firing.
        credential.requires_update = False

        ensure_category_schema()
        sync_summary = {'added': 0, 'modified': 0, 'removed': 0}
        sync_errors = []

        try:
            sync_counts, credential_error = _sync_credential_transactions(credential.user, credential, from_webhook=False)
            for key in sync_summary:
                sync_summary[key] = sync_counts.get(key, 0)
            if credential_error:
                sync_errors.append(credential_error)
        except plaid.ApiException as exc:
            current_app.logger.exception("Initial Plaid sync failed for credential %s: %s", credential.id, exc)
            db.session.rollback()
            return jsonify({
                'status': 'error',
                'message': 'Bank connection created but failed to sync transactions.',
                'error': str(exc)
            }), 502
        except Exception as exc:
            current_app.logger.exception("Unexpected error during initial Plaid sync for credential %s: %s", credential.id, exc)
            db.session.rollback()
            return jsonify({
                'status': 'error',
                'message': 'Bank connection created but failed to sync transactions.',
                'error': str(exc)
            }), 500

        category_summary = None
        has_category_rules = CategoryRule.query.filter_by(user_id=credential.user_id).first() is not None
        if has_category_rules:
            category_summary = apply_category_rules(credential.user_id)

        db.session.commit()
        db.session.refresh(credential)  # Ensure we have the latest cursor state

        response_payload = {
            'status': 'success',
            'message': 'Bank connection added successfully',
            'sync_summary': sync_summary,
            'sync_errors': sync_errors,
            'requires_update': credential.requires_update
        }

        if category_summary is not None:
            response_payload['categories'] = category_summary

        return jsonify(response_payload)

    except Exception as e:
        print(f"Error in handle_token_and_accounts: {str(e)}")
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500


def _persist_transactions_from_payload(user, credential, data, from_webhook=False):
    counts = {'added': 0, 'modified': 0, 'removed': 0}
    payload_by_action = {
        action: list(data.get(action, []) or [])
        for action in ['added', 'modified', 'removed']
    }

    processed_plaid_ids = set()

    transaction_ids = {
        payload.get('transaction_id')
        for payload_list in payload_by_action.values()
        for payload in payload_list
        if payload.get('transaction_id')
    }
    transaction_ids.discard(None)

    existing_transactions = {}
    if transaction_ids:
        existing_transactions = {
            txn.plaid_transaction_id: txn
            for txn in Transaction.query.filter(
                Transaction.plaid_transaction_id.in_(transaction_ids)
            ).all()
        }

    account_map = {account.plaid_account_id: account for account in credential.accounts}
    incoming_account_ids = {
        payload.get('account_id')
        for payload_list in (payload_by_action['added'], payload_by_action['modified'])
        for payload in payload_list
        if payload.get('account_id')
    }
    incoming_account_ids.discard(None)

    missing_account_ids = incoming_account_ids.difference(account_map.keys())
    if missing_account_ids:
        fetched_accounts = Account.query.filter(
            Account.plaid_account_id.in_(missing_account_ids)
        ).all()
        account_map.update({account.plaid_account_id: account for account in fetched_accounts})

    for payload in payload_by_action['removed']:
        plaid_transaction_id = payload.get('transaction_id')
        if not plaid_transaction_id:
            continue

        transaction = existing_transactions.get(plaid_transaction_id)
        if transaction and not transaction.is_removed:
            transaction.is_removed = True
            transaction.last_action = 'removed'
            transaction.updated_at = datetime.utcnow()
            _mark_split_children_removed(transaction)
            counts['removed'] += 1
        elif transaction:
            transaction.last_action = 'removed'
            _mark_split_children_removed(transaction)

    for action in ['added', 'modified']:
        for payload in payload_by_action[action]:
            plaid_transaction_id = payload.get('transaction_id')
            if not plaid_transaction_id or plaid_transaction_id in processed_plaid_ids:
                continue

            processed_plaid_ids.add(plaid_transaction_id)
            is_new_transaction = from_webhook or action == 'added'

            account = account_map.get(payload.get('account_id'))
            if not account:
                current_app.logger.warning(
                    "Skipping transaction %s because account %s was not found",
                    plaid_transaction_id,
                    payload.get('account_id')
                )
                continue

            transaction = Transaction.query.filter_by(plaid_transaction_id=plaid_transaction_id).first()

            if transaction is None:
                transaction = Transaction(
                    plaid_transaction_id=plaid_transaction_id,
                    user_id=user.id,
                    credential_id=credential.id,
                    account_id=account.id,
                    created_at=datetime.utcnow(),
                    is_new=is_new_transaction,
                    seen_by_user=False
                )
                db.session.add(transaction)
            else:
                transaction.credential_id = credential.id
                transaction.account_id = account.id
                if is_new_transaction:
                    transaction.is_new = True
                transaction.seen_by_user = False

            previous_amount = None
            if transaction.id is not None and transaction.amount is not None:
                try:
                    previous_amount = Decimal(transaction.amount)
                except (InvalidOperation, TypeError, ValueError):
                    previous_amount = None

            raw_amount = payload.get('amount', 0)
            if raw_amount is None:
                raw_amount = 0

            transaction.name = payload.get('name') or ''
            try:
                amount_decimal = Decimal(str(raw_amount))
            except (InvalidOperation, TypeError, ValueError):
                amount_decimal = Decimal('0')
            new_amount = -amount_decimal
            transaction.amount = new_amount
            transaction.iso_currency_code = payload.get('iso_currency_code')
            transaction.category = json.dumps(payload.get('category', []))
            transaction.merchant_name = payload.get('merchant_name')
            transaction.payment_channel = payload.get('payment_channel')

            date_value = payload.get('date')
            if isinstance(date_value, datetime):
                transaction.date = date_value.date()
            elif isinstance(date_value, date):
                transaction.date = date_value
            elif isinstance(date_value, str) and date_value:
                try:
                    transaction.date = datetime.strptime(date_value, '%Y-%m-%d').date()
                except ValueError:
                    current_app.logger.warning(
                        "Invalid transaction date %s for %s",
                        date_value,
                        plaid_transaction_id
                    )
            elif date_value:
                current_app.logger.warning(
                    "Unexpected transaction date type %s for %s",
                    type(date_value),
                    plaid_transaction_id
                )
            if transaction.date is None:
                transaction.date = datetime.utcnow().date()

            transaction.pending = bool(payload.get('pending', False))
            transaction.is_removed = False
            transaction.last_action = action
            transaction.updated_at = datetime.utcnow()
            if transaction.has_split_children and previous_amount is not None and previous_amount != new_amount:
                _rebalance_split_children(transaction, previous_amount, new_amount)

            if action in counts:
                counts[action] += 1

    return counts


def _sync_credential_transactions(user, credential, from_webhook=False):
    counts = {'added': 0, 'modified': 0, 'removed': 0}
    cursor = credential.transactions_cursor
    has_more = True
    latest_cursor = cursor
    credential_error = None

    while has_more:
        payload = {
            'access_token': credential.access_token,
            'count': 500
        }

        if cursor:
            payload['cursor'] = cursor

        try:
            response = current_app.plaid_client.transactions_sync(payload)
        except plaid.ApiException as e:
            error_response = json.loads(e.body)
            log_method = current_app.logger.error
            if error_response.get('error_code') == 'ITEM_LOGIN_REQUIRED':
                log_method = current_app.logger.warning

            log_method(
                "Error fetching transactions for credential %s: %s -> %s",
                credential.id,
                error_response.get('error_code'),
                error_response.get('error_message')
            )

            if error_response.get('error_code') == 'DEVELOPMENT_ENVIRONMENT_BROWNOUT':
                raise

            if error_response.get('error_code') == 'ITEM_LOGIN_REQUIRED':
                credential.requires_update = True
                credential_error = {
                    'error_code': error_response['error_code'],
                    'error_message': error_response['error_message']
                }
            else:
                credential_error = {
                    'error_code': error_response.get('error_code'),
                    'error_message': error_response.get('error_message', str(e))
                }
            break
        except Exception as exc:
            current_app.logger.exception("Unexpected error fetching transactions: %s", exc)
            credential_error = {'error_message': str(exc)}
            break

        data = response.to_dict()
        cursor = data.get('next_cursor')
        has_more = data.get('has_more', False)
        latest_cursor = cursor or latest_cursor

        payload_counts = _persist_transactions_from_payload(user, credential, data, from_webhook=from_webhook)
        for key in counts:
            counts[key] += payload_counts.get(key, 0)

    if credential_error is None:
        credential.transactions_cursor = cursor or latest_cursor or credential.transactions_cursor

    return counts, credential_error

@core.route('/api/plaid/webhook', methods=['POST'])
def plaid_webhook():
    data = request.get_json()
    webhook_type = data.get('webhook_type')
    webhook_code = data.get('webhook_code')

    if webhook_type == 'TRANSACTIONS' and webhook_code == 'SYNC_UPDATES_AVAILABLE':
        item_id = data['item_id']
        credential = Credential.query.filter_by(item_id=item_id).first()
        if credential:
            ensure_category_schema()
            try:
                sync_counts, credential_error = _sync_credential_transactions(credential.user, credential, from_webhook=True)
                if credential_error:
                    current_app.logger.warning(
                        "Credential %s reported error after webhook sync: %s",
                        credential.id,
                        credential_error
                    )
                has_category_rules = CategoryRule.query.filter_by(user_id=credential.user_id).first() is not None
                if has_category_rules:
                    apply_category_rules(credential.user_id)
                db.session.commit()
            except Exception as exc:
                current_app.logger.exception(
                    "Failed to process Plaid webhook for credential %s: %s",
                    credential.id,
                    exc
                )
                db.session.rollback()

    return jsonify({'status': 'success'})


@core.route('/api/transactions/sync', methods=['POST'])
@login_required
def sync_transactions():
    ensure_category_schema()
    def handler():
        user = current_user
        active_credentials = Credential.query.filter_by(user_id=user.id, status='Active').all()

        summary = []
        errors = []
        category_summary = None

        for credential in active_credentials:
            try:
                counts, credential_error = _sync_credential_transactions(user, credential, from_webhook=False)
                db.session.commit()

                summary.append({
                    'credential_id': credential.id,
                    'institution_name': credential.institution_name,
                    'added': counts['added'],
                    'modified': counts['modified'],
                    'removed': counts['removed'],
                    'requires_update': credential.requires_update
                })



                if credential_error:
                    errors.append({
                        'credential_id': credential.id,
                        'institution_name': credential.institution_name,
                        **credential_error
                    })
            except plaid.ApiException as e:
                errors.append({
                    'credential_id': credential.id,
                    'institution_name': credential.institution_name,
                    'error_code': 'DEVELOPMENT_ENVIRONMENT_BROWNOUT',
                    'error_message': 'Plaid Development environment is undergoing a scheduled brownout. Please try again later.'
                })
                db.session.rollback()
            except Exception as exc:
                current_app.logger.exception("Failed to sync transactions for credential %s", credential.id)
                errors.append({
                    'credential_id': credential.id,
                    'institution_name': credential.institution_name,
                    'error_message': str(exc)
                })
                db.session.rollback()

        has_category_rules = CategoryRule.query.filter_by(user_id=user.id).first() is not None
        if has_category_rules:
            category_summary = apply_category_rules(user.id)

        db.session.commit()

        status_code = 200 if not errors else 207
        return jsonify({
            'summary': summary,
            'errors': errors,
            'version': VERSION,
            'categories': category_summary
        }), status_code

    return _with_schema_retry(handler)

@core.route('/api/transactions/mark_as_seen', methods=['POST'])
@login_required
def mark_transactions_as_seen():
    data = request.get_json()
    transaction_ids = data.get('transaction_ids', [])

    if not transaction_ids:
        return jsonify({'error': 'No transaction_ids provided'}), 400

    transactions = Transaction.query.filter(Transaction.id.in_(transaction_ids), Transaction.user_id == current_user.id).all()

    for txn in transactions:
        txn.seen_by_user = True
        txn.is_new = False
        txn.last_seen_by_user = datetime.utcnow()

    db.session.commit()

    return jsonify({'status': 'success'})

@core.route('/api/accounts', methods=['GET'])
@login_required
def get_accounts():
    user_id = current_user.id
    bank_id = request.args.get('bank_id')

    query = Account.query.join(Credential).filter(Credential.user_id == user_id, Credential.status == 'Active')

    if bank_id:
        query = query.filter(Credential.id == bank_id)

    accounts = query.filter(Account.status == 'Active').all()

    accounts_data = [{'id': account.id, 'name': account.name, 'mask': account.mask, 
                    'type': account.type, 'subtype': account.subtype, 'is_enabled': account.is_enabled} for account in accounts]

    return jsonify(accounts=accounts_data)

@core.route('/api/banks', methods=['GET'])
@login_required
def get_banks():
    banks = Credential.query.filter_by(user_id=current_user.id, status='Active').all()
    banks_data = []

    for bank in banks:
        accounts_data = []
        for account in bank.accounts:
            if account.status != 'Active':
                continue

            accounts_data.append({
                'id': account.id,
                'name': account.name,
                'mask': account.mask,
                'type': account.type,
                'subtype': account.subtype,
                'plaid_account_id': account.plaid_account_id,
                'is_enabled': account.is_enabled
            })

        banks_data.append({
            'id': bank.id,
            'institution_name': bank.institution_name,
            'requires_update': bank.requires_update,
            'accounts': accounts_data
        })

    return jsonify(banks=banks_data)

@core.route('/api/transactions', methods=['GET'])
@login_required
def get_transactions():
    ensure_category_schema()

    def handler():
        filter_options = _parse_transaction_filters(request.args)
        account_ids = filter_options['account_ids']
        page = filter_options['page']
        page_size = filter_options['page_size']
        sort_key = filter_options['sort_key']
        sort_desc = filter_options['sort_desc']
        search_term = filter_options['search_term']
        start_date = filter_options['start_date']
        end_date = filter_options['end_date']
        min_amount = filter_options['min_amount']
        max_amount = filter_options['max_amount']
        custom_category_param = filter_options['custom_category_param']

        query = _build_transactions_query(current_user.id, filter_options)

        total_count = query.count()

        sort_mapping = {
            'date': Transaction.date,
            'name': Transaction.name,
            'amount': Transaction.amount
        }
        sort_column = sort_mapping.get(sort_key, Transaction.date)
        primary_order = desc(sort_column) if sort_desc else asc(sort_column)

        secondary_order = desc(Transaction.id)
        if sort_column is Transaction.date and not sort_desc:
            secondary_order = asc(Transaction.id)

        ordered_query = query.order_by(primary_order, secondary_order)
        offset = (page - 1) * page_size
        transactions = ordered_query.offset(offset).limit(page_size).all()

        current_app.logger.info(
            'Transaction fetch filters start=%s end=%s min=%s max=%s custom=%s page=%s size=%s total=%s',
            start_date.isoformat() if start_date else None,
            end_date.isoformat() if end_date else None,
            str(min_amount) if min_amount is not None else None,
            str(max_amount) if max_amount is not None else None,
            custom_category_param or None,
            page,
            page_size,
            total_count
        )
        if transactions:
            current_app.logger.info(
                'Transaction fetch sample first_date=%s last_date=%s',
                transactions[0].date.isoformat() if transactions[0].date else None,
                transactions[-1].date.isoformat() if transactions[-1].date else None
            )

        override_map = _load_overrides([txn.id for txn in transactions])
        response = [_serialize_transaction(txn, override_map) for txn in transactions]

        has_more = (page * page_size) < total_count

        return jsonify(
            transactions=response,
            page=page,
            page_size=page_size,
            total_count=total_count,
            has_more=has_more,
            sort_key=sort_key,
            sort_desc=sort_desc,
            search=search_term,
            filters={
                'start_date': start_date.isoformat() if start_date else None,
                'end_date': end_date.isoformat() if end_date else None,
                'min_amount': str(min_amount) if isinstance(min_amount, Decimal) else (str(min_amount) if min_amount not in (None, '') else None),
                'max_amount': str(max_amount) if isinstance(max_amount, Decimal) else (str(max_amount) if max_amount not in (None, '') else None),
                'custom_category_id': custom_category_param or None
            }
        )

    return _with_schema_retry(handler)


@core.route('/api/transactions/export', methods=['GET'])
@login_required
def export_transactions():
    ensure_category_schema()

    def handler():
        format_param = (request.args.get('format') or 'csv').strip().lower()
        if format_param not in ('csv', 'xlsx'):
            return jsonify({'error': 'Unsupported export format.'}), 400

        if format_param == 'xlsx' and Workbook is None:
            return jsonify({'error': 'Excel export is not available on this server.'}), 501

        filters = _parse_transaction_filters(request.args)
        query = _build_transactions_query(current_user.id, filters)

        sort_mapping = {
            'date': Transaction.date,
            'name': Transaction.name,
            'amount': Transaction.amount
        }
        sort_column = sort_mapping.get(filters['sort_key'], Transaction.date)
        primary_order = desc(sort_column) if filters['sort_desc'] else asc(sort_column)
        secondary_order = desc(Transaction.id)
        if sort_column is Transaction.date and not filters['sort_desc']:
            secondary_order = asc(Transaction.id)

        transactions = query.order_by(primary_order, secondary_order).all()
        override_map = _load_overrides([txn.id for txn in transactions])
        payload = [_serialize_transaction(txn, override_map) for txn in transactions]

        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        if format_param == 'csv':
            csv_content = _export_transactions_to_csv(payload)
            filename = f'transactions_{timestamp}.csv'
            response = current_app.response_class(csv_content, mimetype='text/csv')
            response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        output_stream = _export_transactions_to_excel(payload)
        filename = f'transactions_{timestamp}.xlsx'
        output_stream.seek(0)
        return send_file(
            output_stream,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    return _with_schema_retry(handler)


@core.route('/api/transactions/<int:transaction_id>/split', methods=['GET', 'POST'])
@login_required
def split_transaction(transaction_id):
    ensure_category_schema()

    def handler():
        txn = Transaction.query.options(
            joinedload(Transaction.account),
            joinedload(Transaction.credential),
            joinedload(Transaction.custom_category)
        ).filter_by(id=transaction_id, user_id=current_user.id).first()

        if not txn or txn.is_removed:
            return jsonify({'error': 'Transaction not found.'}), 404

        if txn.is_split_child:
            return jsonify({'error': 'Split transactions cannot be split again.'}), 400

        if request.method == 'GET':
            child_query = txn.split_children.filter(
                Transaction.is_removed.is_(False),
                Transaction.is_split_child.is_(True)
            ).order_by(Transaction.id.asc())
            children = child_query.all()
            override_map = _load_overrides([txn.id] + [child.id for child in children])
            return jsonify({
                'parent': _serialize_transaction(txn, override_map),
                'children': [_serialize_transaction(child, override_map) for child in children]
            })

        payload = request.get_json() or {}
        splits = payload.get('splits')
        if not isinstance(splits, list):
            return jsonify({'error': 'Invalid payload.'}), 400

        if len(splits) < 2:
            return jsonify({'error': 'Provide at least two split rows.'}), 400

        parent_amount = Decimal(txn.amount or 0)
        parent_abs = parent_amount.copy_abs().quantize(CENT, rounding=ROUND_HALF_UP)
        if parent_abs <= Decimal('0.00'):
            return jsonify({'error': 'Only transactions with a non-zero amount can be split.'}), 400

        plaid_categories = _extract_category_list(txn.category)
        fallback_label = ' / '.join(plaid_categories) if plaid_categories else None
        plaid_candidates = list(plaid_categories)
        if fallback_label:
            plaid_candidates.append(fallback_label)

        seen_categories = set()
        processed = []
        for index, split in enumerate(splits, start=1):
            description = (split.get('description') or '').strip()
            category_label = (split.get('category') or '').strip()
            if not description:
                return jsonify({'error': f'Row {index}: Description is required.'}), 400
            if not category_label:
                return jsonify({'error': f'Row {index}: Category is required.'}), 400

            try:
                cleaned_category = _validate_category_name(category_label, plaid_candidates=plaid_candidates)
            except ValueError as exc:
                return jsonify({'error': f'Row {index}: {exc}'}), 400

            normalized_category = cleaned_category.lower()
            if normalized_category in seen_categories:
                return jsonify({'error': 'Each split must use a unique category.'}), 400
            seen_categories.add(normalized_category)

            try:
                amount_abs = _normalize_split_amount(split.get('amount'))
            except ValueError as exc:
                return jsonify({'error': f'Row {index}: {exc}'}), 400

            processed.append({
                'description': description,
                'category': cleaned_category,
                'amount': amount_abs
            })

        total_amount = sum((item['amount'] for item in processed), Decimal('0.00'))
        remainder = (parent_abs - total_amount).quantize(CENT, rounding=ROUND_HALF_UP)

        if abs(remainder) > CENT:
            return jsonify({'error': 'Split amounts must total the transaction amount.'}), 400

        if processed:
            adjusted = (processed[-1]['amount'] + remainder).quantize(CENT, rounding=ROUND_HALF_UP)
            if adjusted <= Decimal('0.00'):
                return jsonify({'error': 'Each split amount must be greater than zero.'}), 400
            processed[-1]['amount'] = adjusted

        children = _create_split_children(txn, processed)

        db.session.flush()
        override_map = _load_overrides([txn.id] + [child.id for child in children])

        response_payload = {
            'parent': _serialize_transaction(txn, override_map),
            'children': [_serialize_transaction(child, override_map) for child in children]
        }
        db.session.commit()
        return jsonify(response_payload), 201

    return _with_schema_retry(handler)


@core.route('/api/transactions/<int:transaction_id>/category', methods=['PATCH'])
@login_required
def update_transaction_category(transaction_id):
    ensure_category_schema()

    def handler():
        txn = Transaction.query.options(
            joinedload(Transaction.account),
            joinedload(Transaction.credential),
            joinedload(Transaction.custom_category)
        ).filter_by(id=transaction_id, user_id=current_user.id).first()

        if not txn:
            return jsonify({'error': 'Transaction not found.'}), 404

        payload = request.get_json() or {}
        raw_label = payload.get('label')
        label = _normalize_category_name(raw_label)
        explicit_color = payload.get('color')
        force_create = payload.get('force_create', False)

        override = TransactionCategoryOverride.query.filter_by(transaction_id=txn.id).first()

        if not label:
            if override:
                db.session.delete(override)
            txn.custom_category_id = None
            db.session.flush()
            apply_rules_to_transactions(current_user.id, transaction_ids=[txn.id])
            db.session.commit()
            override_map = _load_overrides([txn.id])
            return jsonify({'transaction': _serialize_transaction(txn, override_map)})

        plaid_categories = _extract_category_list(txn.category)
        fallback_label = ' / '.join(plaid_categories) if plaid_categories else None
        plaid_candidates = []
        if fallback_label:
            plaid_candidates.append(fallback_label)

        existing_category = _find_custom_category(current_user.id, label)
        normalized_color = None
        if explicit_color:
            try:
                normalized_color = _normalize_color(explicit_color)
            except ValueError:
                normalized_color = None

        if existing_category:
            category = existing_category
            if normalized_color and category.color != normalized_color:
                category.color = normalized_color
        else:
            trimmed = _normalize_category_name(label)
            if len(trimmed) < 3:
                return jsonify({'error': 'Category name must be at least 3 characters.'}), 400

            if not force_create:
                lowered = trimmed.lower()
                for candidate in plaid_candidates:
                    if not candidate:
                        continue
                    candidate_trimmed = _normalize_category_name(candidate)
                    if not candidate_trimmed:
                        continue
                    if lowered == candidate_trimmed.lower():
                        return jsonify({
                            'confirmation_required': True,
                            'message': 'An Automatic category with this name already exists. Do you want to create a new custom category with the same name?'
                        }), 409

            category = CustomCategory(
                user_id=current_user.id,
                name=trimmed,
                color=normalized_color or DEFAULT_MANUAL_COLOR
            )
            db.session.add(category)
            db.session.flush()

        if override is None:
            override = TransactionCategoryOverride(transaction_id=txn.id, custom_category_id=category.id)
            db.session.add(override)
        else:
            override.custom_category_id = category.id

        txn.custom_category_id = None

        db.session.flush()
        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            current_app.logger.warning('Failed to assign custom category "%s" for user %s due to integrity error', label, current_user.id, exc_info=True)
            return jsonify({'error': 'Could not save the category. Please try again or choose a different name.'}), 400

        override_map = _load_overrides([txn.id])
        return jsonify({'transaction': _serialize_transaction(txn, override_map)})

    return _with_schema_retry(handler)


@core.route('/api/custom-categories', methods=['GET'])
@login_required
def list_custom_categories():
    ensure_category_schema()

    def handler():
        categories = CustomCategory.query.filter_by(user_id=current_user.id).order_by(
            func.lower(CustomCategory.name).asc(),
            CustomCategory.id.asc()
        ).all()
        category_ids = [category.id for category in categories]
        rule_counts, transaction_counts, override_counts = _collect_custom_category_stats(current_user.id, category_ids)
        response = [
            _serialize_custom_category(
                category,
                {
                    'rule_count': int(rule_counts.get(category.id, 0)),
                    'transaction_count': int(transaction_counts.get(category.id, 0)),
                    'override_count': int(override_counts.get(category.id, 0))
                }
            )
            for category in categories
        ]
        return jsonify({
            'categories': response,
            'labels': _collect_category_labels(current_user.id)
        })

    return _with_schema_retry(handler)


@core.route('/api/custom-categories', methods=['POST'])
@login_required
def create_custom_category():
    ensure_category_schema()

    def handler():
        payload = request.get_json() or {}
        raw_name = payload.get('name') or payload.get('label')
        try:
            name = _validate_category_name(raw_name, plaid_candidates=None)
        except ValueError as exc:
            return jsonify({'error': str(exc)}), 400
        if len(name) > 255:
            name = name[:255]
        if not name:
            return jsonify({'error': 'Name is required.'}), 400

        color_value = payload.get('color') or DEFAULT_MANUAL_COLOR
        try:
            color = _normalize_color(color_value)
        except ValueError as exc:
            return jsonify({'error': str(exc)}), 400

        existing = _find_custom_category(current_user.id, name)
        if existing:
            return jsonify({'error': 'A category with this name already exists.'}), 400

        category = CustomCategory(
            user_id=current_user.id,
            name=name,
            color=color
        )
        db.session.add(category)
        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            current_app.logger.warning('Duplicate custom category creation attempt for user %s and name "%s"', current_user.id, name)
            return jsonify({'error': 'A category with this name already exists.'}), 409

        payload_category = _serialize_custom_category(category, {
            'rule_count': 0,
            'transaction_count': 0,
            'override_count': 0
        })
        return jsonify({
            'category': payload_category,
            'labels': _collect_category_labels(current_user.id)
        }), 201

    return _with_schema_retry(handler)


@core.route('/api/custom-categories/<int:category_id>', methods=['PUT'])
@login_required
def update_custom_category(category_id):
    ensure_category_schema()

    def handler():
        category = CustomCategory.query.filter_by(id=category_id, user_id=current_user.id).first()
        if not category:
            return jsonify({'error': 'Custom category not found.'}), 404

        payload = request.get_json() or {}
        name = payload.get('name')
        color_value = payload.get('color')
        updated = False

        if name is not None:
            try:
                trimmed = _validate_category_name(name, plaid_candidates=None)
            except ValueError as exc:
                return jsonify({'error': str(exc)}), 400
            if len(trimmed) > 255:
                trimmed = trimmed[:255]
            if trimmed.lower() != (category.name or '').lower():
                duplicate = CustomCategory.query.filter(
                    CustomCategory.user_id == current_user.id,
                    func.lower(CustomCategory.name) == trimmed.lower(),
                    CustomCategory.id != category.id
                ).first()
                if duplicate:
                    return jsonify({'error': 'Another category with this name already exists.'}), 400
                category.name = trimmed
                updated = True

        if color_value is not None:
            try:
                normalized_color = _normalize_color(color_value)
            except ValueError as exc:
                return jsonify({'error': str(exc)}), 400
            if category.color != normalized_color:
                category.color = normalized_color
                updated = True

        if updated:
            db.session.flush()
        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            current_app.logger.warning('Duplicate custom category update attempt for user %s and name "%s"', current_user.id, name)
            return jsonify({'error': 'Another category with this name already exists.'}), 409

        rule_counts, transaction_counts, override_counts = _collect_custom_category_stats(current_user.id, [category.id])
        payload_category = _serialize_custom_category(
            category,
            {
                'rule_count': int(rule_counts.get(category.id, 0)),
                'transaction_count': int(transaction_counts.get(category.id, 0)),
                'override_count': int(override_counts.get(category.id, 0))
            }
        )
        return jsonify({
            'category': payload_category,
            'labels': _collect_category_labels(current_user.id)
        })

    return _with_schema_retry(handler)


@core.route('/api/custom-categories/<int:category_id>', methods=['DELETE'])
@login_required
def delete_custom_category(category_id):
    ensure_category_schema()

    def handler():
        category = CustomCategory.query.filter_by(id=category_id, user_id=current_user.id).first()
        if not category:
            return jsonify({'error': 'Custom category not found.'}), 404

        overrides = TransactionCategoryOverride.query.join(Transaction).filter(
            TransactionCategoryOverride.custom_category_id == category.id,
            Transaction.user_id == current_user.id
        ).all()
        for override in overrides:
            db.session.delete(override)

        Transaction.query.filter(
            Transaction.user_id == current_user.id,
            Transaction.custom_category_id == category.id
        ).update({Transaction.custom_category_id: None}, synchronize_session=False)

        CategoryRule.query.filter_by(user_id=current_user.id, category_id=category.id).delete(synchronize_session=False)

        db.session.delete(category)
        db.session.flush()

        summary = apply_category_rules(current_user.id)
        db.session.commit()

        return jsonify({
            'deleted': category_id,
            'summary': summary,
            'labels': _collect_category_labels(current_user.id)
        })

    return _with_schema_retry(handler)


@core.route('/api/categories', methods=['GET'])
@login_required
def list_categories():
    ensure_category_schema()

    def handler():
        rules = CategoryRule.query.options(
            joinedload(CategoryRule.category)
        ).filter_by(user_id=current_user.id).order_by(CategoryRule.created_at.desc(), CategoryRule.id.desc()).all()
        return jsonify({
            'categories': [_serialize_category(rule) for rule in rules],
            'labels': _collect_category_labels(current_user.id)
        })

    return _with_schema_retry(handler)


@core.route('/api/categories', methods=['POST'])
@login_required
def create_category_rule():
    ensure_category_schema()
    def handler():
        payload = request.get_json() or {}
        text_to_match = (payload.get('text_to_match') or '').strip()
        category_id = payload.get('category_id')
        category_name = (_extract_label(payload) or '').strip()

        if not text_to_match:
            return jsonify({'error': 'Text to match is required.'}), 400
        if not category_id and not category_name:
            return jsonify({'error': 'A category selection is required.'}), 400

        if len(text_to_match) > 512:
            text_to_match = text_to_match[:512]
        if category_name and len(category_name) > 255:
            category_name = category_name[:255]

        field_to_match = _resolve_rule_field(payload.get('field_to_match'))
        transaction_type = _resolve_rule_type(payload.get('transaction_type'))

        try:
            amount_min = _parse_decimal_value(payload.get('amount_min'))
            amount_max = _parse_decimal_value(payload.get('amount_max'))
        except ValueError as exc:
            return jsonify({'error': str(exc)}), 400

        if amount_min is not None and amount_max is not None and amount_min > amount_max:
            return jsonify({'error': 'Amount >= cannot be greater than Amount <='}), 400

        category = None
        if category_id:
            try:
                category_id = int(category_id)
            except (TypeError, ValueError):
                return jsonify({'error': 'Invalid category id.'}), 400
            category = CustomCategory.query.filter_by(id=category_id, user_id=current_user.id).first()
            if not category:
                return jsonify({'error': 'Selected category was not found.'}), 404
        else:
            try:
                category = _get_or_create_custom_category(
                    current_user.id,
                    category_name,
                    color=payload.get('color')
                )
            except ValueError as exc:
                return jsonify({'error': str(exc)}), 400

        rule = CategoryRule(
            user_id=current_user.id,
            category_id=category.id,
            text_to_match=text_to_match,
            field_to_match=field_to_match,
            transaction_type=transaction_type,
            amount_min=amount_min,
            amount_max=amount_max
        )
        rule.category = category
        db.session.add(rule)
        db.session.flush()

        summary = apply_category_rules(current_user.id)
        db.session.commit()

        return jsonify({
            'category': _serialize_category(rule),
            'summary': summary,
            'labels': _collect_category_labels(current_user.id)
        }), 201

    return _with_schema_retry(handler)


@core.route('/api/categories/<int:category_id>', methods=['PUT'])
@login_required
def update_category_rule(category_id):
    ensure_category_schema()
    def handler():
        rule = CategoryRule.query.options(
            joinedload(CategoryRule.category)
        ).filter_by(id=category_id, user_id=current_user.id).first()
        if not rule:
            return jsonify({'error': 'Category rule not found.'}), 404

        payload = request.get_json() or {}
        text_to_match = (payload.get('text_to_match') or '').strip()
        category_id_payload = payload.get('category_id')
        category_name = (_extract_label(payload) or '').strip()

        if not text_to_match:
            return jsonify({'error': 'Text to match is required.'}), 400
        if not category_id_payload and not category_name:
            return jsonify({'error': 'A category selection is required.'}), 400

        if len(text_to_match) > 512:
            text_to_match = text_to_match[:512]
        if category_name and len(category_name) > 255:
            category_name = category_name[:255]

        field_to_match = _resolve_rule_field(payload.get('field_to_match'))
        transaction_type = _resolve_rule_type(payload.get('transaction_type'))

        try:
            amount_min = _parse_decimal_value(payload.get('amount_min'))
            amount_max = _parse_decimal_value(payload.get('amount_max'))
        except ValueError as exc:
            return jsonify({'error': str(exc)}), 400

        if amount_min is not None and amount_max is not None and amount_min > amount_max:
            return jsonify({'error': 'Amount >= cannot be greater than Amount <='}), 400

        category = None
        if category_id_payload:
            try:
                category_id_value = int(category_id_payload)
            except (TypeError, ValueError):
                return jsonify({'error': 'Invalid category id.'}), 400
            category = CustomCategory.query.filter_by(id=category_id_value, user_id=current_user.id).first()
            if not category:
                return jsonify({'error': 'Selected category was not found.'}), 404
        else:
            try:
                category = _get_or_create_custom_category(
                    current_user.id,
                    category_name,
                    color=payload.get('color')
                )
            except ValueError as exc:
                return jsonify({'error': str(exc)}), 400

        rule.text_to_match = text_to_match
        rule.field_to_match = field_to_match
        rule.transaction_type = transaction_type
        rule.amount_min = amount_min
        rule.amount_max = amount_max
        rule.category_id = category.id
        rule.category = category

        db.session.flush()

        summary = apply_category_rules(current_user.id)
        db.session.commit()

        return jsonify({
            'category': _serialize_category(rule),
            'summary': summary,
            'labels': _collect_category_labels(current_user.id)
        })

    return _with_schema_retry(handler)


@core.route('/api/categories/<int:category_id>', methods=['DELETE'])
@login_required
def delete_category_rule(category_id):
    ensure_category_schema()
    def handler():
        rule = CategoryRule.query.filter_by(id=category_id, user_id=current_user.id).first()
        if not rule:
            return jsonify({'error': 'Category rule not found.'}), 404

        db.session.delete(rule)
        db.session.flush()

        summary = apply_category_rules(current_user.id)
        db.session.commit()

        return jsonify({
            'deleted': category_id,
            'summary': summary,
            'labels': _collect_category_labels(current_user.id)
        })

    return _with_schema_retry(handler)


@core.route('/api/dashboard', methods=['GET'])
@login_required
def dashboard_summary():
    ensure_category_schema()
    balances = _collect_balances_summary(current_user.id)
    spending = _collect_spending_summary(current_user.id)
    cashflow = _collect_cashflow_summary(current_user.id)
    return jsonify({
        'balances': balances,
        'spending': spending,
        'cashflow': cashflow
    })


@core.route('/api/budgets', methods=['GET'])
@login_required
def list_budgets():
    budgets = Budget.query.filter_by(user_id=current_user.id).order_by(Budget.created_at.asc(), Budget.id.asc()).all()
    metrics = _calculate_budget_metrics(current_user.id, [budget.category_label for budget in budgets])
    response = [_serialize_budget(budget, metrics) for budget in budgets]
    summary = _build_budget_summary(budgets, metrics)
    return jsonify({'budgets': response, 'summary': summary})


@core.route('/api/budgets', methods=['POST'])
@login_required
def create_budget():
    payload = request.get_json() or {}
    category_label = (payload.get('category_label') or '').strip()
    if not category_label:
        return jsonify({'error': 'Category is required.'}), 400

    try:
        frequency = _normalize_budget_frequency(payload.get('frequency'))
        amount = _normalize_budget_amount(payload.get('amount'))
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400

    budget = Budget(
        user_id=current_user.id,
        category_label=category_label,
        frequency=frequency,
        amount=amount
    )
    db.session.add(budget)
    db.session.commit()

    budgets = Budget.query.filter_by(user_id=current_user.id).order_by(Budget.created_at.asc(), Budget.id.asc()).all()
    metrics = _calculate_budget_metrics(current_user.id, [item.category_label for item in budgets])
    response_budget = _serialize_budget(budget, metrics)
    summary = _build_budget_summary(budgets, metrics)

    return jsonify({'budget': response_budget, 'summary': summary}), 201


@core.route('/api/budgets/<int:budget_id>', methods=['PUT'])
@login_required
def update_budget(budget_id):
    budget = Budget.query.filter_by(id=budget_id, user_id=current_user.id).first()
    if not budget:
        return jsonify({'error': 'Budget not found.'}), 404

    payload = request.get_json() or {}
    category_label = (payload.get('category_label') or '').strip()
    if not category_label:
        return jsonify({'error': 'Category is required.'}), 400

    try:
        frequency = _normalize_budget_frequency(payload.get('frequency'))
        amount = _normalize_budget_amount(payload.get('amount'))
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400

    budget.category_label = category_label
    budget.frequency = frequency
    budget.amount = amount
    db.session.commit()

    budgets = Budget.query.filter_by(user_id=current_user.id).order_by(Budget.created_at.asc(), Budget.id.asc()).all()
    metrics = _calculate_budget_metrics(current_user.id, [item.category_label for item in budgets])
    response_budget = _serialize_budget(budget, metrics)
    summary = _build_budget_summary(budgets, metrics)

    return jsonify({'budget': response_budget, 'summary': summary})


@core.route('/api/budgets/<int:budget_id>', methods=['DELETE'])
@login_required
def delete_budget(budget_id):
    budget = Budget.query.filter_by(id=budget_id, user_id=current_user.id).first()
    if not budget:
        return jsonify({'error': 'Budget not found.'}), 404

    db.session.delete(budget)
    db.session.commit()

    budgets = Budget.query.filter_by(user_id=current_user.id).order_by(Budget.created_at.asc(), Budget.id.asc()).all()
    metrics = _calculate_budget_metrics(current_user.id, [item.category_label for item in budgets])
    summary = _build_budget_summary(budgets, metrics)
    remaining = [_serialize_budget(item, metrics) for item in budgets]

    return jsonify({'budgets': remaining, 'summary': summary})


@core.route('/api/balance', methods=['POST'])
@login_required
def fetch_balances():
    data = request.json
    access_token = data.get('access_token')
    account_ids = data.get('account_ids', [])

    # Use AccountsGetRequest instead of AccountsBalanceGetRequest
    accounts_request = AccountsGetRequest(
        access_token=access_token,
        options={"account_ids": account_ids} if account_ids else None
    )

    try:
        accounts_response = current_app.plaid_client.accounts_get(accounts_request)
        accounts = accounts_response.to_dict().get('accounts', [])

        # Extract the necessary balance information from the accounts data
        balances = []
        for account in accounts:
            balance_info = {
                "account_id": account.get('account_id'),
                "balances": account.get('balances')
            }
            balances.append(balance_info)

        return jsonify({"balances": balances})
    except plaid.ApiException as e:
        return jsonify(json.loads(e.body)), e.status

@core.route('/api/remove_bank/<int:bank_id>', methods=['DELETE'])
@login_required
def remove_bank(bank_id):
    credential = Credential.query.filter_by(id=bank_id, user_id=current_user.id).first()
    
    if credential:
        success, plaid_response = deactivate_plaid_token(credential.access_token)
        
        if success:
            credential.status = 'Revoked'

            for account in credential.accounts:
                account.status = 'Revoked'
            

            db.session.commit()
            session['connections_modal_open'] = True
            return jsonify({'success': True, 'message': 'Bank connection removed'}), 200
        else:
            print(f"Failed to deactivate token: {plaid_response}")
            session['connections_modal_open'] = True
            return jsonify({'success': False, 'message': 'Failed to remove bank connection', 'error': plaid_response}), 400
    else:
        return jsonify({'success': False, 'message': 'Credential not found'}), 404

@core.route('/api/get_access_token/<int:credential_id>', methods=['GET'])
@login_required
def get_access_token_for_bank(credential_id):
    credential = Credential.query.filter_by(id=credential_id, user_id=current_user.id).first()

    if not credential:
        return jsonify({'error': 'Credential not found or does not belong to the current user'}), 404

    try:
        access_token = credential.access_token
        return jsonify({'access_token': access_token})
    except Exception as e:
        return jsonify({'error': 'Failed to retrieve access token', 'message': str(e)}), 500

# CANDIDATE FOR DELETION, THIS FUNCTION DOES NOT SEEM TO BE USED ANYWHERE:
def refresh_accounts(credential_id, accounts_data):
    existing_accounts = Account.query.filter_by(credential_id=credential_id).all()
    existing_account_ids = {account.plaid_account_id for account in existing_accounts}

    for account in accounts_data['accounts']:
        if account['account_id'] in existing_account_ids:
            existing_account = next((acc for acc in existing_accounts if acc.plaid_account_id == account['account_id']), None)
            if existing_account:
                existing_account.name = account['name']
                existing_account.type = account['type']
                existing_account.subtype = account['subtype']
                existing_account.mask = account.get('mask', '')
        else:
                new_account = Account(
                    credential_id=credential_id,
                    plaid_account_id=account['account_id'],
                    name=account['name'],
                    type=account['type'],
                    subtype=account['subtype'],
                    mask=account.get('mask', ''),
                    status='Active'
                )
                db.session.add(new_account)

    current_account_ids = {account['account_id'] for account in accounts_data['accounts']}
    for existing_account in existing_accounts:
        if existing_account.plaid_account_id not in current_account_ids:
            existing_account.status = 'Inactive'

    db.session.commit()

@core.route('/plaid_webhook', methods=['POST'])
def plaid_webhook_item_error():
    data = request.get_json()

    # Extract relevant fields from the payload
    timestamp = datetime.now()
    error = data.get('error', {}) or {}  # Handle None by defaulting to an empty dictionary
    error_code = error.get('error_code', 'None')
    item_id = data.get('item_id', 'Unknown')
    webhook_code = data.get('webhook_code', 'Unknown')
    webhook_type = data.get('webhook_type', 'Unknown')

    # Log the incoming webhook
    print("***** INCOMING WEBHOOK *****")
    print(f"Data: {data}")

    # Handle ITEM_LOGIN_REQUIRED
    if webhook_type == "ITEM" and webhook_code in ["ERROR", "NEW_ACCOUNTS_AVAILABLE"]:
        try:
            # Find the related Credential in the database
            credential = Credential.query.filter_by(item_id=item_id).first()

            if credential:
                # Extract the user_id from the Credential
                user_id = credential.user_id
                if not user_id:
                    print(f"Warning: No user_id found for Credential ID: {credential.id}")
                    return jsonify({"status": "ok"}), 200

                credential.requires_update = True
                db.session.commit()
                print(f"Updated Credential (ID: {credential.id}) with requires_update=True")


            else:
                print(f"No Credential found for Item ID: {item_id}")
        except Exception as e:
            print(f"Error handling ITEM_LOGIN_REQUIRED: {str(e)}")
            db.session.rollback()

    return jsonify({"status": "ok"}), 200


def _format_transaction_categories(transaction):
    categories = transaction.get('category') if isinstance(transaction, dict) else None
    if not categories:
        return ''
    if isinstance(categories, list):
        return ' / '.join([str(item) for item in categories if item])
    return str(categories)


def _coerce_transaction_date(value):
    if not value:
        return ''
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(value)
    except (ValueError, TypeError):
        return value


def _export_transactions_to_csv(transactions):
    headers = [
        'Date',
        'Description',
        'Merchant',
        'Amount',
        'Currency',
        'Category',
        'Automatic Categories',
        'Account',
        'Account Mask',
        'Institution',
        'Pending',
        'Transaction ID'
    ]
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)

    for txn in transactions:
        amount_value = txn.get('amount')
        try:
            amount_decimal = Decimal(str(amount_value)).quantize(CENT, rounding=ROUND_HALF_UP)
        except (InvalidOperation, TypeError, ValueError):
            amount_decimal = Decimal('0.00')
        pending_label = 'Yes' if txn.get('pending') else 'No'
        writer.writerow([
            txn.get('date') or '',
            txn.get('name') or '',
            txn.get('merchant_name') or '',
            str(amount_decimal),
            txn.get('iso_currency_code') or '',
            txn.get('custom_category') or '',
            _format_transaction_categories(txn),
            txn.get('account_name') or '',
            txn.get('account_mask') or '',
            txn.get('institution_name') or '',
            pending_label,
            txn.get('plaid_transaction_id') or ''
        ])

    return output.getvalue()


def _export_transactions_to_excel(transactions):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Transactions'

    headers = [
        'Date',
        'Description',
        'Merchant',
        'Amount',
        'Currency',
        'Category',
        'Automatic Categories',
        'Account',
        'Account Mask',
        'Institution',
        'Pending',
        'Transaction ID'
    ]
    worksheet.append(headers)

    for txn in transactions:
        amount_value = txn.get('amount')
        try:
            amount_decimal = Decimal(str(amount_value)).quantize(CENT, rounding=ROUND_HALF_UP)
        except (InvalidOperation, TypeError, ValueError):
            amount_decimal = Decimal('0.00')
        pending_label = 'Yes' if txn.get('pending') else 'No'
        worksheet.append([
            _coerce_transaction_date(txn.get('date')),
            txn.get('name') or '',
            txn.get('merchant_name') or '',
            float(amount_decimal),
            txn.get('iso_currency_code') or '',
            txn.get('custom_category') or '',
            _format_transaction_categories(txn),
            txn.get('account_name') or '',
            txn.get('account_mask') or '',
            txn.get('institution_name') or '',
            pending_label,
            txn.get('plaid_transaction_id') or ''
        ])

    if Workbook is not None and 'worksheet' in locals() and 'workbook' in locals():
        try:
            from openpyxl.utils import get_column_letter  # Lazy import for optional dependency
            for idx, column_cells in enumerate(worksheet.columns, 1):
                max_length = 0
                for cell in column_cells:
                    value = cell.value
                    if isinstance(value, date):
                        text = value.isoformat()
                    elif value is None:
                        text = ''
                    else:
                        text = str(value)
                    if len(text) > max_length:
                        max_length = len(text)
                worksheet.column_dimensions[get_column_letter(idx)].width = min(40, max(10, max_length + 2))
        except Exception:
            pass

    output = BytesIO()
    workbook.save(output)
    return output


def json_csv(transactions, action):
    si = StringIO()
    cw = csv.writer(si)

    for item in transactions:
        row = [
            item.get("date", ""),
            item.get("name", ""),
            item.get("amount", ""),
            item.get("iso_currency_code", ""),
            ", ".join(item.get("category", [])),
            item.get("merchant_name", ""),
            item.get("account_id", ""),
            item.get("transaction_id", ""),
            item.get("payment_channel", ""),
            action,
            item.get("pending", "")
        ]
        cw.writerow(row)

    return si.getvalue()

def filter_transactions(transactions, account_plaid_id):
    return [transaction for transaction in transactions if transaction['account_id'] == account_plaid_id]

def deactivate_plaid_token(access_token):
    try:
        print('Access Token to Remove: ', access_token)
        request = ItemRemoveRequest(access_token=access_token)
        response = current_app.plaid_client.item_remove(request)
        
        print("Plaid response:", response)
        return True, response.to_dict()
    except plaid.ApiException as e:
        print("An error occurred while removing the item from Plaid:", e)
        return False, e.body